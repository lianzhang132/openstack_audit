"""
Flask主应用
"""
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response
from datetime import datetime, timedelta
import json
import logging
from io import BytesIO
from uuid import uuid4

from config import Config
from models import db, init_db, Project, VirtualMachine, ChangeLog, AttachedResource
from openstack_service import openstack_service
from scheduler import init_scheduler, sync_openstack_data

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config.from_object(Config)

# 初始化数据库
init_db(app)

# 初始化定时任务
if app.config['ENABLE_SCHEDULER']:
    init_scheduler(app)


# ============ 模板全局函数 ============
@app.context_processor
def utility_processor():
    """注入模板全局函数"""
    return {
        'now': datetime.now
    }

# ============ 首页 ============
@app.route('/')
def index():
    """首页 - 仪表盘"""
    # 先更新过期状态
    update_expire_status()

    stats = {
        'total_vms': VirtualMachine.query.count(),
        'in_use': VirtualMachine.query.filter_by(status=VirtualMachine.STATUS_IN_USE).count(),
        'vm_pending': VirtualMachine.query.filter_by(status=VirtualMachine.STATUS_VM_PENDING).count(),
        'attached_pending': VirtualMachine.query.filter_by(status=VirtualMachine.STATUS_ATTACHED_PENDING).count(),
        'recycled': VirtualMachine.query.filter_by(status=VirtualMachine.STATUS_RECYCLED).count(),
        'long_term': VirtualMachine.query.filter_by(is_long_term=True).count(),
        'projects': Project.query.count(),
    }

    # 即将过期的VM（7天内）
    today = datetime.now().date()
    expire_soon = VirtualMachine.query.filter(
        VirtualMachine.status == VirtualMachine.STATUS_IN_USE,
        VirtualMachine.is_long_term == False,
        VirtualMachine.expire_date <= today + timedelta(days=7),
        VirtualMachine.expire_date >= today
    ).order_by(VirtualMachine.expire_date).limit(10).all()

    # 最近变更记录
    recent_logs = ChangeLog.query.order_by(ChangeLog.changed_at.desc()).limit(10).all()

    return render_template('index.html', stats=stats, expire_soon=expire_soon, recent_logs=recent_logs)


def update_expire_status():
    """更新过期状态"""
    today = datetime.now().date()

    # 1. 查找已过期但状态还是"使用期内"的VM → 进入"虚机待回收"
    expired_vms = VirtualMachine.query.filter(
        VirtualMachine.status == VirtualMachine.STATUS_IN_USE,
        VirtualMachine.is_long_term == False,
        VirtualMachine.expire_date != None,
        VirtualMachine.expire_date < today
    ).all()

    for vm in expired_vms:
        vm.status = VirtualMachine.STATUS_VM_PENDING

        log = ChangeLog(
            vm_id=vm.id,
            change_type=ChangeLog.TYPE_STATUS_CHANGE,
            operator='system'
        )
        log.set_change_content({
            'message': '云主机已过期，进入虚机待回收状态',
            'expire_date': vm.expire_date.strftime('%Y-%m-%d') if vm.expire_date else None,
            'old_status': VirtualMachine.STATUS_IN_USE,
            'new_status': VirtualMachine.STATUS_VM_PENDING
        })
        db.session.add(log)

    # 2. 查找虚机已删除且附属资源已过期的 → 进入"附属待回收"
    #    需要排除：附属资源长期保留、附属资源未过期的
    pending_vms = VirtualMachine.query.filter(
        VirtualMachine.status == VirtualMachine.STATUS_VM_PENDING,
        VirtualMachine.vm_deleted_date != None
    ).all()

    for vm in pending_vms:
        # 检查附属资源是否过期
        if vm.is_attached_expired():
            vm.status = VirtualMachine.STATUS_ATTACHED_PENDING

            # 更新附属资源状态
            for resource in vm.attached_resources.filter_by(status=AttachedResource.STATUS_ACTIVE):
                resource.status = AttachedResource.STATUS_PENDING
                resource.pending_at = datetime.now()

            log = ChangeLog(
                vm_id=vm.id,
                change_type=ChangeLog.TYPE_STATUS_CHANGE,
                operator='system'
            )
            expire_date = vm.get_attached_expire_date()
            log.set_change_content({
                'message': '附属资源已过期，进入待回收状态',
                'attached_expire_date': expire_date.strftime('%Y-%m-%d') if expire_date else None,
                'old_status': VirtualMachine.STATUS_VM_PENDING,
                'new_status': VirtualMachine.STATUS_ATTACHED_PENDING
            })
            db.session.add(log)

    if expired_vms or any(vm.status == VirtualMachine.STATUS_ATTACHED_PENDING for vm in pending_vms):
        db.session.commit()


def _parse_date(value):
    if value:
        return datetime.strptime(value, '%Y-%m-%d').date()
    return None


def _get_resource_id(resource):
    if isinstance(resource, dict):
        return resource.get('id')
    return getattr(resource, 'id', None)


def _is_truthy(value):
    return str(value).strip().lower() in ('1', 'true', 'yes', 'on')


def apply_business_form_fields(vm, form, default_status=None):
    """将页面上的业务审计字段写入VM记录"""
    vm.purpose = form.get('purpose', '')
    vm.is_long_term = form.get('is_long_term') == 'on'
    vm.long_term_reason = form.get('long_term_reason', '')
    vm.applicant = form.get('applicant', '')

    status = form.get('status', default_status or VirtualMachine.STATUS_IN_USE)
    if status in [s[0] for s in VirtualMachine.STATUS_CHOICES]:
        vm.status = status

    vm.created_date = _parse_date(form.get('created_date')) or datetime.now().date()
    expire_date = _parse_date(form.get('expire_date'))
    if expire_date:
        vm.expire_date = expire_date
    elif not vm.is_long_term:
        vm.expire_date = vm.created_date + timedelta(days=Config.VM_DEFAULT_EXPIRE_DAYS)
    else:
        vm.expire_date = None

    vm_deleted_date = _parse_date(form.get('vm_deleted_date'))
    if vm_deleted_date:
        vm.vm_deleted_date = vm_deleted_date

    project_id = form.get('project_id')
    if project_id:
        vm.project_id = int(project_id)


def update_vm_record_from_detail(vm, detail):
    """将OpenStack详情写入VM记录"""
    vm.name = detail.get('name')
    vm.host = detail.get('host')
    vm.flavor = detail.get('flavor')
    vm.openstack_status = detail.get('status')
    vm.openstack_project_id = detail.get('project_id')

    if detail.get('image'):
        vm.image_info = json.dumps(detail['image'], ensure_ascii=False)
    if detail.get('flavor_detail'):
        vm.flavor_detail = json.dumps(detail['flavor_detail'], ensure_ascii=False)

    vm.set_networks(detail.get('networks', []))
    vm.set_volumes(detail.get('volumes', []))
    vm.last_sync_at = datetime.now()
    vm.last_sync_data = json.dumps(detail, ensure_ascii=False, default=str)


def append_attached_resources_from_detail(vm, detail):
    """根据OpenStack详情创建附属资源审计记录"""
    for port in detail.get('networks', []):
        resource_id = port.get('port_id') or port.get('id') or port.get('mac')
        if resource_id:
            resource = AttachedResource(
                resource_type=AttachedResource.TYPE_PORT,
                resource_id=resource_id,
                resource_name=port.get('port_name') or port.get('network_name'),
            )
            resource.set_resource_info(port)
            vm.attached_resources.append(resource)

    for vol in detail.get('volumes', []):
        vol_id = vol.get('id')
        if vol_id:
            resource = AttachedResource(
                resource_type=AttachedResource.TYPE_VOLUME,
                resource_id=vol_id,
                resource_name=vol.get('name'),
            )
            resource.set_resource_info(vol)
            vm.attached_resources.append(resource)


def has_attached_resource(vm, resource_type, resource_id):
    if not resource_id:
        return False
    return vm.attached_resources.filter_by(
        resource_type=resource_type,
        resource_id=resource_id
    ).first() is not None


def ensure_attached_resource(vm, resource_type, resource_id, resource_name=None, resource_info=None):
    if not resource_id or has_attached_resource(vm, resource_type, resource_id):
        return
    resource = AttachedResource(
        resource_type=resource_type,
        resource_id=resource_id,
        resource_name=resource_name,
    )
    if resource_info:
        resource.set_resource_info(resource_info)
    vm.attached_resources.append(resource)


def build_network_from_created_port(port_info, network_id):
    fixed_ips = port_info.get('fixed_ips') or []
    fixed_ip = fixed_ips[0].get('ip_address') if fixed_ips else ''
    return {
        'network_id': network_id,
        'network_name': port_info.get('network_name') or network_id,
        'port_id': port_info.get('id'),
        'port_name': port_info.get('name'),
        'ip': fixed_ip,
        'mac': port_info.get('mac'),
        'port_security_enabled': port_info.get('port_security_enabled'),
    }


def build_fallback_detail(server_id, form, created_port_info=None):
    boot_source = form.get('boot_source', 'image')
    detail = {
        'uuid': server_id,
        'name': form.get('create_name', '').strip(),
        'status': 'BUILD',
        'project_id': None,
        'host': None,
        'image': None,
        'flavor': form.get('flavor_id'),
        'flavor_detail': {},
        'networks': [],
        'volumes': [],
    }

    if boot_source == 'image':
        detail['image'] = {'id': form.get('image_id'), 'name': form.get('image_id')}
    else:
        detail['volumes'].append({
            'id': form.get('boot_volume_id'),
            'name': form.get('boot_volume_id'),
            'bootable': True,
        })

    if created_port_info:
        detail['networks'].append(build_network_from_created_port(created_port_info, form.get('network_id')))
    else:
        detail['networks'].append({
            'network_id': form.get('network_id'),
            'network_name': form.get('network_id'),
            'ip': '',
            'mac': '',
        })

    return detail


def create_failed_create_record(form, created_port_info, error_message):
    """端口已创建但虚机创建失败时，保留可审计、可清理的本地记录"""
    vm = VirtualMachine(
        uuid=f"create-failed-{uuid4().hex}",
        name=form.get('create_name', '').strip() or 'create-failed',
        status=VirtualMachine.STATUS_VM_PENDING,
        openstack_status='CREATE_FAILED',
        vm_deleted_date=datetime.now().date(),
    )
    apply_business_form_fields(vm, form, default_status=VirtualMachine.STATUS_VM_PENDING)
    vm.status = VirtualMachine.STATUS_VM_PENDING
    vm.openstack_status = 'CREATE_FAILED'
    vm.vm_deleted_date = datetime.now().date()

    if created_port_info:
        network = build_network_from_created_port(created_port_info, form.get('network_id'))
        vm.set_networks([network])
        resource = AttachedResource(
            resource_type=AttachedResource.TYPE_PORT,
            resource_id=created_port_info.get('id') or created_port_info.get('mac'),
            resource_name=created_port_info.get('name') or form.get('port_name') or form.get('network_id'),
        )
        resource.set_resource_info(network)
        vm.attached_resources.append(resource)

    db.session.add(vm)
    db.session.flush()

    if created_port_info:
        port_log = ChangeLog(
            vm_id=vm.id,
            change_type=ChangeLog.TYPE_CREATE,
            operator=form.get('operator') or form.get('applicant') or 'admin'
        )
        port_log.set_change_content({
            'message': '创建网络端口',
            'action': 'create_port',
            'port': created_port_info,
            'disable_port_security': created_port_info.get('port_security_enabled') is False,
            'server_create_failed': True,
        })
        db.session.add(port_log)

    log = ChangeLog(
        vm_id=vm.id,
        change_type=ChangeLog.TYPE_CREATE,
        operator=form.get('operator') or form.get('applicant') or 'admin'
    )
    log.set_change_content({
        'message': '创建云主机失败，已保留待清理记录',
        'action': 'create_server_failed',
        'error': error_message,
        'created_port': created_port_info,
    })
    db.session.add(log)
    db.session.commit()
    return vm


def handle_openstack_vm_create():
    form = request.form
    name = form.get('create_name', '').strip()
    boot_source = form.get('boot_source', 'image')
    flavor_id = form.get('flavor_id', '').strip()
    image_id = form.get('image_id', '').strip()
    boot_volume_id = form.get('boot_volume_id', '').strip()
    network_id = form.get('network_id', '').strip()
    create_port = form.get('create_port') == 'on'
    disable_port_security = form.get('disable_port_security') == 'on'
    created_port_info = None

    if not name:
        flash('云主机名称不能为空', 'error')
        return redirect(url_for('vm_add'))
    if not flavor_id:
        flash('请选择规格', 'error')
        return redirect(url_for('vm_add'))
    if boot_source == 'image' and not image_id:
        flash('镜像启动需要选择镜像', 'error')
        return redirect(url_for('vm_add'))
    if boot_source == 'volume' and not boot_volume_id:
        flash('卷启动需要选择启动卷', 'error')
        return redirect(url_for('vm_add'))
    if not network_id:
        flash('请选择网络', 'error')
        return redirect(url_for('vm_add'))

    try:
        if create_port:
            created_port_info = openstack_service.create_port(
                network_id=network_id,
                name=form.get('port_name', '').strip() or f"{name}-port",
                fixed_ip=form.get('fixed_ip', '').strip() or None,
                disable_port_security=disable_port_security,
            )

        server = openstack_service.create_server(
            name=name,
            flavor_id=flavor_id,
            image_id=image_id if boot_source == 'image' else None,
            boot_volume_id=boot_volume_id if boot_source == 'volume' else None,
            network_id=network_id if not created_port_info else None,
            port_id=created_port_info.get('id') if created_port_info else None,
            delete_volume_on_termination=form.get('delete_volume_on_termination') == 'on',
        )
        server_id = _get_resource_id(server)
        if not server_id:
            raise RuntimeError('OpenStack未返回云主机ID')

        if VirtualMachine.query.filter_by(uuid=server_id).first():
            raise RuntimeError(f'云主机记录已存在: {server_id}')

        detail = openstack_service.get_server_detail(server_id)
        if not detail:
            detail = build_fallback_detail(server_id, form, created_port_info)

        vm = VirtualMachine(uuid=server_id)
        update_vm_record_from_detail(vm, detail)
        apply_business_form_fields(vm, form, default_status=VirtualMachine.STATUS_IN_USE)
        db.session.add(vm)
        db.session.flush()
        append_attached_resources_from_detail(vm, detail)

        if created_port_info:
            port_network = build_network_from_created_port(created_port_info, network_id)
            ensure_attached_resource(
                vm,
                AttachedResource.TYPE_PORT,
                created_port_info.get('id') or created_port_info.get('mac'),
                created_port_info.get('name') or form.get('port_name') or network_id,
                port_network,
            )
            if not any(net.get('port_id') == created_port_info.get('id') for net in vm.get_networks()):
                networks = vm.get_networks()
                networks.append(port_network)
                vm.set_networks(networks)

        if boot_source == 'volume':
            volume_info = {
                'id': boot_volume_id,
                'name': boot_volume_id,
                'bootable': True,
                'delete_on_termination': form.get('delete_volume_on_termination') == 'on',
            }
            ensure_attached_resource(
                vm,
                AttachedResource.TYPE_VOLUME,
                boot_volume_id,
                boot_volume_id,
                volume_info,
            )
            if not any(volume.get('id') == boot_volume_id for volume in vm.get_volumes()):
                volumes = vm.get_volumes()
                volumes.append(volume_info)
                vm.set_volumes(volumes)

        if created_port_info:
            port_log = ChangeLog(
                vm_id=vm.id,
                change_type=ChangeLog.TYPE_CREATE,
                operator=form.get('operator') or form.get('applicant') or 'admin'
            )
            port_log.set_change_content({
                'message': '创建网络端口',
                'action': 'create_port',
                'port': created_port_info,
                'disable_port_security': disable_port_security,
            })
            db.session.add(port_log)

        create_log = ChangeLog(
            vm_id=vm.id,
            change_type=ChangeLog.TYPE_CREATE,
            operator=form.get('operator') or form.get('applicant') or 'admin'
        )
        create_log.set_change_content({
            'message': '创建云主机',
            'action': 'create_server',
            'server_id': server_id,
            'boot_source': boot_source,
            'image_id': image_id if boot_source == 'image' else None,
            'boot_volume_id': boot_volume_id if boot_source == 'volume' else None,
            'flavor_id': flavor_id,
            'network_id': network_id,
            'created_port_id': created_port_info.get('id') if created_port_info else None,
        })
        db.session.add(create_log)
        db.session.commit()

        flash('云主机创建成功，已写入审计记录', 'success')
        return redirect(url_for('vm_detail', id=vm.id))

    except Exception as e:
        db.session.rollback()
        if created_port_info:
            try:
                failed_vm = create_failed_create_record(form, created_port_info, str(e))
                flash(f'云主机创建失败，已记录已创建端口以便后续回收: {str(e)}', 'error')
                return redirect(url_for('vm_detail', id=failed_vm.id))
            except Exception as audit_err:
                db.session.rollback()
                logger.error(f"Failed to create audit record for failed VM create: {audit_err}")
        flash(f'云主机创建失败: {str(e)}', 'error')
        return redirect(url_for('vm_add'))

# ============ Bug 3 修复：Excel导出 ============
def to_export_dict(self):
    """导出用的字典"""
    fd = self.get_flavor_detail()
    attached_expire = self.get_attached_expire_date()

    return {
        '云主机名称': self.name or '-',
        'UUID': self.uuid,
        '所属项目': self.project.name if self.project else '-',
        '用途说明': self.purpose or '-',
        '申请人': self.applicant or '-',
        '是否长期使用': '是' if self.is_long_term else '否',
        '长期使用理由': self.long_term_reason or '-',
        '镜像': self.get_image_display(),
        '规格': self.flavor or '-',
        'CPU(核)': fd.get('vcpus', '-') if fd else '-',
        '内存(MB)': fd.get('ram', '-') if fd else '-',
        '系统盘(GB)': fd.get('disk', '-') if fd else '-',
        '网络信息': self.get_networks_display(),
        '存储卷': self.get_volumes_display(),
        '宿主机': self.host or '-',
        '创建日期': self.created_date.strftime('%Y-%m-%d') if self.created_date else '-',
        '到期日期': self.expire_date.strftime('%Y-%m-%d') if self.expire_date else (
            '-' if not self.is_long_term else '长期'),
        '当前状态': self.get_status_display(),
        'OpenStack状态': self.openstack_status or '-',
        '虚机删除日期': self.vm_deleted_date.strftime('%Y-%m-%d') if self.vm_deleted_date else '-',
        '附属资源长期保留': '是' if self.is_attached_long_term else '否',
        '附属资源长期理由': self.attached_long_term_reason or '-',
        '附属资源到期日期': attached_expire.strftime('%Y-%m-%d') if attached_expire else (
            '-' if not self.is_attached_long_term else '长期'),
        'OpenStack项目': self.openstack_project_name or self.openstack_project_id or '-',
        '记录创建时间': self.record_created_at.strftime('%Y-%m-%d %H:%M:%S') if self.record_created_at else '-',
        '最后同步时间': self.last_sync_at.strftime('%Y-%m-%d %H:%M:%S') if self.last_sync_at else '-',
    }


# ============ Bug 1 修复：同步和初始化接口改为POST，返回后重定向 ============
@app.route('/api/sync/all', methods=['POST'])
def api_sync_all():
    """手动触发全量同步"""
    try:
        sync_openstack_data(app)
        flash('同步完成', 'success')
    except Exception as e:
        flash(f'同步失败: {str(e)}', 'error')

    return redirect(url_for('index'))


@app.route('/vms/export')
def vm_export():
    """导出云主机列表为Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('导出功能需要安装openpyxl库', 'error')
        return redirect(url_for('vm_list'))

    # 获取筛选参数
    status = request.args.get('status', '')
    project_id = request.args.get('project_id', '')
    search = request.args.get('search', '')
    is_long_term = request.args.get('is_long_term', '')

    query = VirtualMachine.query

    if status:
        query = query.filter_by(status=status)
    if project_id:
        query = query.filter_by(project_id=int(project_id))
    if is_long_term:
        query = query.filter_by(is_long_term=(is_long_term == 'true'))
    if search:
        query = query.filter(
            (VirtualMachine.name.contains(search)) |
            (VirtualMachine.uuid.contains(search)) |
            (VirtualMachine.applicant.contains(search)) |
            (VirtualMachine.networks.contains(search)) |
            (VirtualMachine.purpose.contains(search))
        )

    vms = query.order_by(VirtualMachine.record_created_at.desc()).all()

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "云主机跟踪记录表"

    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头（添加附属资源相关列）
    headers = [
        '序号', '云主机名称', 'UUID', '所属项目', '用途说明', '申请人',
        '是否长期使用', '长期使用理由', '镜像', '规格', 'CPU(核)', '内存(MB)',
        '系统盘(GB)', '网络信息', '存储卷', '宿主机', '创建日期', '到期日期',
        '当前状态', 'OpenStack状态', '虚机删除日期',
        '附属资源长期保留', '附属资源长期理由', '附属资源到期日期',
        'OpenStack项目', '记录创建时间', '最后同步时间'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据行
    for row_idx, vm in enumerate(vms, 2):
        data = vm.to_export_dict()
        row_data = [
            row_idx - 1,
            data['云主机名称'],
            data['UUID'],
            data['所属项目'],
            data['用途说明'],
            data['申请人'],
            data['是否长期使用'],
            data['长期使用理由'],
            data['镜像'],
            data['规格'],
            data['CPU(核)'],
            data['内存(MB)'],
            data['系统盘(GB)'],
            data['网络信息'],
            data['存储卷'],
            data['宿主机'],
            data['创建日期'],
            data['到期日期'],
            data['当前状态'],
            data['OpenStack状态'],
            data['虚机删除日期'],
            data['附属资源长期保留'],
            data['附属资源长期理由'],
            data['附属资源到期日期'],
            data['OpenStack项目'],
            data['记录创建时间'],
            data['最后同步时间'],
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # 调整列宽（添加新列的宽度）
    column_widths = [
        6,  # 序号
        20,  # 云主机名称
        36,  # UUID
        15,  # 所属项目
        30,  # 用途说明
        10,  # 申请人
        12,  # 是否长期使用
        25,  # 长期使用理由
        20,  # 镜像
        15,  # 规格
        8,  # CPU
        10,  # 内存
        10,  # 系统盘
        30,  # 网络信息
        25,  # 存储卷
        15,  # 宿主机
        12,  # 创建日期
        12,  # 到期日期
        12,  # 当前状态
        12,  # OpenStack状态
        12,  # 虚机删除日期
        15,  # 附属资源长期保留
        20,  # 附属资源长期理由
        15,  # 附属资源到期日期
        15,  # OpenStack项目
        20,  # 记录创建时间
        20,  # 最后同步时间
    ]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    from urllib.parse import quote
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename_cn = f"云主机跟踪记录表_{timestamp}.xlsx"
    filename_encoded = quote(filename_cn)

    response = Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{filename_encoded}"

    return response

@app.route('/api/vms/init', methods=['POST'])
def api_init_vms():
    """初始化：从OpenStack导入所有VM"""
    try:
        servers = openstack_service.list_servers(all_projects=True)
        count = 0

        for server in servers:
            existing = VirtualMachine.query.filter_by(uuid=server.id).first()
            if existing:
                continue

            detail = openstack_service.get_server_detail(server.id)
            if not detail:
                continue

            vm = VirtualMachine(uuid=server.id)
            vm.name = detail.get('name')
            vm.host = detail.get('host')
            vm.flavor = detail.get('flavor')
            vm.openstack_status = detail.get('status')
            vm.openstack_project_id = detail.get('project_id')

            if detail.get('image'):
                vm.image_info = json.dumps(detail['image'], ensure_ascii=False)
            if detail.get('flavor_detail'):
                vm.flavor_detail = json.dumps(detail['flavor_detail'], ensure_ascii=False)

            vm.set_networks(detail.get('networks', []))
            vm.set_volumes(detail.get('volumes', []))

            vm.created_date = datetime.now().date()
            vm.expire_date = datetime.now().date() + timedelta(days=Config.VM_DEFAULT_EXPIRE_DAYS)
            vm.last_sync_at = datetime.now()
            vm.last_sync_data = json.dumps(detail, ensure_ascii=False, default=str)

            db.session.add(vm)
            db.session.flush()

            log = ChangeLog(
                vm_id=vm.id,
                change_type=ChangeLog.TYPE_CREATE,
                operator='system_init'
            )
            log.set_change_content({'message': '系统初始化导入'})
            db.session.add(log)

            count += 1

        db.session.commit()
        flash(f'成功导入 {count} 个云主机', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'导入失败: {str(e)}', 'error')

    return redirect(url_for('index'))


# ============ 项目管理（公司内部项目） ============
@app.route('/projects')
def project_list():
    """项目列表"""
    projects = Project.query.order_by(Project.created_at.desc()).all()
    return render_template('projects.html', projects=projects)


@app.route('/projects/add', methods=['GET', 'POST'])
def project_add():
    """添加项目"""
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('项目名称不能为空', 'error')
            return redirect(url_for('project_add'))

        # 检查是否已存在
        existing = Project.query.filter_by(name=name).first()
        if existing:
            flash('项目名称已存在', 'error')
            return redirect(url_for('project_add'))

        project = Project(
            name=name,
            description=request.form.get('description', ''),
            owner=request.form.get('owner', ''),
            department=request.form.get('department', '')
        )
        db.session.add(project)
        db.session.commit()

        flash('项目添加成功', 'success')
        return redirect(url_for('project_list'))

    return render_template('project_add.html')


@app.route('/projects/<int:id>/edit', methods=['GET', 'POST'])
def project_edit(id):
    """编辑项目"""
    project = Project.query.get_or_404(id)

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('项目名称不能为空', 'error')
            return redirect(url_for('project_edit', id=id))

        # 检查名称是否与其他项目重复
        existing = Project.query.filter(Project.name == name, Project.id != id).first()
        if existing:
            flash('项目名称已存在', 'error')
            return redirect(url_for('project_edit', id=id))

        project.name = name
        project.description = request.form.get('description', '')
        project.owner = request.form.get('owner', '')
        project.department = request.form.get('department', '')

        db.session.commit()
        flash('项目更新成功', 'success')
        return redirect(url_for('project_list'))

    return render_template('project_edit.html', project=project)


@app.route('/projects/<int:id>/delete', methods=['POST'])
def project_delete(id):
    """删除项目"""
    project = Project.query.get_or_404(id)

    # 检查是否有关联的VM
    if project.virtual_machines.count() > 0:
        flash('该项目下还有云主机记录，无法删除', 'error')
        return redirect(url_for('project_list'))

    db.session.delete(project)
    db.session.commit()
    flash('项目删除成功', 'success')
    return redirect(url_for('project_list'))


# ============ 云主机管理 ============
@app.route('/vms')
def vm_list():
    """云主机列表"""
    # 筛选参数
    status = request.args.get('status', '')
    project_id = request.args.get('project_id', '')
    search = request.args.get('search', '')
    is_long_term = request.args.get('is_long_term', '')

    query = VirtualMachine.query

    if status:
        query = query.filter_by(status=status)
    if project_id:
        query = query.filter_by(project_id=int(project_id))
    if is_long_term:
        query = query.filter_by(is_long_term=(is_long_term == 'true'))
    if search:
        query = query.filter(
            (VirtualMachine.name.contains(search)) |
            (VirtualMachine.uuid.contains(search)) |
            (VirtualMachine.applicant.contains(search)) |
            (VirtualMachine.networks.contains(search)) |
            (VirtualMachine.purpose.contains(search))
        )

    vms = query.order_by(VirtualMachine.record_created_at.desc()).all()
    projects = Project.query.order_by(Project.name).all()

    return render_template('vm_list.html', vms=vms, projects=projects,
                          current_status=status, current_project=project_id,
                          search=search, current_long_term=is_long_term)




@app.route('/vms/add', methods=['GET', 'POST'])
def vm_add():
    """添加云主机记录"""
    if request.method == 'POST':
        add_mode = request.form.get('add_mode', 'auto')  # auto 或 manual
        if add_mode == 'create':
            return handle_openstack_vm_create()

        uuid = request.form.get('uuid', '').strip()

        if not uuid:
            flash('UUID不能为空', 'error')
            return redirect(url_for('vm_add'))

        # 检查是否已存在
        existing = VirtualMachine.query.filter_by(uuid=uuid).first()
        if existing:
            flash('该云主机记录已存在', 'error')
            return redirect(url_for('vm_add'))

        vm = VirtualMachine(uuid=uuid)

        # 根据模式获取OpenStack信息
        if add_mode == 'auto':
            # 自动模式：从OpenStack获取信息
            detail = openstack_service.get_server_detail(uuid)

            if detail:
                vm.name = detail.get('name')
                vm.host = detail.get('host')
                vm.flavor = detail.get('flavor')
                vm.openstack_status = detail.get('status')
                vm.openstack_project_id = detail.get('project_id')

                if detail.get('image'):
                    vm.image_info = json.dumps(detail['image'], ensure_ascii=False)
                if detail.get('flavor_detail'):
                    vm.flavor_detail = json.dumps(detail['flavor_detail'], ensure_ascii=False)

                vm.set_networks(detail.get('networks', []))
                vm.set_volumes(detail.get('volumes', []))

                vm.last_sync_at = datetime.now()
                vm.last_sync_data = json.dumps(detail, ensure_ascii=False, default=str)

                # 创建附属资源记录
                for port in detail.get('networks', []):
                    resource_id = port.get('port_id') or port.get('id') or port.get('mac')
                    if resource_id:
                        resource = AttachedResource(
                            resource_type=AttachedResource.TYPE_PORT,
                            resource_id=resource_id,
                            resource_name=port.get('port_name') or port.get('network_name'),
                        )
                        resource.set_resource_info(port)
                        vm.attached_resources.append(resource)

                for vol in detail.get('volumes', []):
                    resource = AttachedResource(
                        resource_type=AttachedResource.TYPE_VOLUME,
                        resource_id=vol.get('id'),
                        resource_name=vol.get('name'),
                    )
                    resource.set_resource_info(vol)
                    vm.attached_resources.append(resource)
            else:
                flash('未能从OpenStack获取信息，请使用手动模式添加', 'warning')
        else:
            # 手动模式：手动填写所有信息
            vm.name = request.form.get('name', '').strip() or None
            vm.host = request.form.get('host', '').strip() or None
            vm.flavor = request.form.get('flavor', '').strip() or None
            vm.openstack_status = request.form.get('openstack_status', 'DELETED')

            # 镜像信息
            image_name = request.form.get('image_name', '').strip()
            if image_name:
                vm.image_info = json.dumps({'name': image_name}, ensure_ascii=False)

            # 规格详情
            vcpus = request.form.get('vcpus', '').strip()
            ram = request.form.get('ram', '').strip()
            disk = request.form.get('disk', '').strip()
            if vcpus or ram or disk:
                flavor_detail = {}
                if vcpus:
                    flavor_detail['vcpus'] = int(vcpus)
                if ram:
                    flavor_detail['ram'] = int(ram)
                if disk:
                    flavor_detail['disk'] = int(disk)
                vm.flavor_detail = json.dumps(flavor_detail, ensure_ascii=False)

            # 网络信息
            ip_addresses = request.form.get('ip_addresses', '').strip()
            if ip_addresses:
                networks = []
                for ip in ip_addresses.split('\n'):
                    ip = ip.strip()
                    if ip:
                        networks.append({'ip': ip, 'network_name': '', 'mac': ''})
                vm.set_networks(networks)

            # 存储卷信息
            volume_info = request.form.get('volume_info', '').strip()
            if volume_info:
                volumes = []
                for vol in volume_info.split('\n'):
                    vol = vol.strip()
                    if vol:
                        volumes.append({'name': vol, 'id': '', 'size': ''})
                vm.set_volumes(volumes)

        # 手动输入字段（两种模式都需要）
        vm.purpose = request.form.get('purpose', '')
        vm.is_long_term = request.form.get('is_long_term') == 'on'
        vm.long_term_reason = request.form.get('long_term_reason', '')
        vm.applicant = request.form.get('applicant', '')

        # 状态
        status = request.form.get('status', VirtualMachine.STATUS_IN_USE)
        if status in [s[0] for s in VirtualMachine.STATUS_CHOICES]:
            vm.status = status

        # 创建日期
        created_date_str = request.form.get('created_date')
        if created_date_str:
            vm.created_date = datetime.strptime(created_date_str, '%Y-%m-%d').date()
        else:
            vm.created_date = datetime.now().date()

        # 设置到期日期
        expire_date_str = request.form.get('expire_date')
        if expire_date_str:
            vm.expire_date = datetime.strptime(expire_date_str, '%Y-%m-%d').date()
        elif not vm.is_long_term:
            vm.expire_date = vm.created_date + timedelta(days=Config.VM_DEFAULT_EXPIRE_DAYS)

        # VM删除日期（用于历史记录）
        vm_deleted_date_str = request.form.get('vm_deleted_date')
        if vm_deleted_date_str:
            vm.vm_deleted_date = datetime.strptime(vm_deleted_date_str, '%Y-%m-%d').date()

        # 项目关联
        project_id = request.form.get('project_id')
        if project_id:
            vm.project_id = int(project_id)

        db.session.add(vm)
        db.session.flush()

        # 创建变更记录
        log = ChangeLog(
            vm_id=vm.id,
            change_type=ChangeLog.TYPE_CREATE,
            operator=request.form.get('applicant', 'admin')
        )
        log.set_change_content({
            'message': '创建云主机记录',
            'mode': '手动添加' if add_mode == 'manual' else '自动获取'
        })
        db.session.add(log)

        db.session.commit()

        flash('云主机记录添加成功', 'success')
        return redirect(url_for('vm_detail', id=vm.id))

    # GET请求
    projects = Project.query.order_by(Project.name).all()

    # 从OpenStack获取可选的服务器列表
    available_servers = []
    images = []
    flavors = []
    networks = []
    volumes = []
    try:
        servers = openstack_service.list_servers(all_projects=True)
        existing_uuids = {vm.uuid for vm in VirtualMachine.query.all()}
        available_servers = [s for s in servers if s.id not in existing_uuids]
    except:
        pass

    try:
        images = openstack_service.list_images()
        flavors = openstack_service.list_flavors()
        networks = openstack_service.list_networks()
        volumes = [
            volume for volume in openstack_service.list_volumes(all_projects=True)
            if str(volume.get('status', '')).lower() == 'available' and _is_truthy(volume.get('bootable'))
        ]
    except Exception as e:
        logger.warning(f"Failed to load OpenStack create resources: {e}")

    return render_template(
        'vm_add.html',
        projects=projects,
        available_servers=available_servers,
        images=images,
        flavors=flavors,
        networks=networks,
        volumes=volumes,
    )


@app.route('/vms/<int:id>')
def vm_detail(id):
    """云主机详情"""
    vm = VirtualMachine.query.get_or_404(id)
    logs = vm.change_logs.order_by(ChangeLog.changed_at.desc()).all()
    resources = vm.attached_resources.all()
    projects = Project.query.order_by(Project.name).all()

    return render_template('vm_detail.html', vm=vm, logs=logs, resources=resources, projects=projects)


@app.route('/vms/<int:id>/edit', methods=['GET', 'POST'])
def vm_edit(id):
    """编辑云主机"""
    vm = VirtualMachine.query.get_or_404(id)

    if request.method == 'POST':
        changes = []

        # 项目变更
        new_project_id = request.form.get('project_id')
        new_project_id = int(new_project_id) if new_project_id else None
        if vm.project_id != new_project_id:
            old_project = vm.project.name if vm.project else None
            new_project = Project.query.get(new_project_id) if new_project_id else None
            changes.append({
                'field': '所属项目',
                'old': old_project,
                'new': new_project.name if new_project else None
            })
            vm.project_id = new_project_id

        # 用途变更
        new_purpose = request.form.get('purpose', '')
        if vm.purpose != new_purpose:
            changes.append({'field': '用途说明', 'old': vm.purpose, 'new': new_purpose})
            vm.purpose = new_purpose

        # 长期使用变更
        new_long_term = request.form.get('is_long_term') == 'on'
        if vm.is_long_term != new_long_term:
            changes.append({'field': '是否长期使用', 'old': '是' if vm.is_long_term else '否', 'new': '是' if new_long_term else '否'})
            vm.is_long_term = new_long_term
            # 如果改为非长期，重新计算到期日期
            if not new_long_term and vm.created_date:
                vm.expire_date = vm.created_date + timedelta(days=Config.VM_DEFAULT_EXPIRE_DAYS)
            elif new_long_term:
                vm.expire_date = None

        # 长期理由变更
        new_reason = request.form.get('long_term_reason', '')
        if vm.long_term_reason != new_reason:
            changes.append({'field': '长期使用理由', 'old': vm.long_term_reason, 'new': new_reason})
            vm.long_term_reason = new_reason

        # 申请人变更
        new_applicant = request.form.get('applicant', '')
        if vm.applicant != new_applicant:
            changes.append({'field': '申请人', 'old': vm.applicant, 'new': new_applicant})
            vm.applicant = new_applicant

        # 创建日期变更
        created_date_str = request.form.get('created_date')
        if created_date_str:
            new_date = datetime.strptime(created_date_str, '%Y-%m-%d').date()
            if vm.created_date != new_date:
                changes.append({
                    'field': '创建日期',
                    'old': vm.created_date.strftime('%Y-%m-%d') if vm.created_date else None,
                    'new': new_date.strftime('%Y-%m-%d')
                })
                vm.created_date = new_date
                if not vm.is_long_term:
                    vm.expire_date = new_date + timedelta(days=Config.VM_DEFAULT_EXPIRE_DAYS)

        if changes:
            log = ChangeLog(
                vm_id=vm.id,
                change_type=ChangeLog.TYPE_UPDATE,
                operator=request.form.get('operator', 'admin')
            )
            log.set_change_content({'message': '手动更新信息', 'changes': changes})
            db.session.add(log)

        db.session.commit()
        flash('更新成功', 'success')
        return redirect(url_for('vm_detail', id=vm.id))

    projects = Project.query.order_by(Project.name).all()
    return render_template('vm_edit.html', vm=vm, projects=projects)


@app.route('/vms/<int:id>/sync')
def vm_sync(id):
    """同步单个VM信息"""
    vm = VirtualMachine.query.get_or_404(id)

    try:
        detail = openstack_service.get_server_detail(vm.uuid)

        if detail:
            from scheduler import detect_changes, update_vm_from_detail
            changes = detect_changes(vm, detail)

            if changes:
                log = ChangeLog(
                    vm_id=vm.id,
                    change_type=ChangeLog.TYPE_SYNC,
                    operator='manual_sync'
                )
                log.set_change_content({'message': '手动同步', 'changes': changes})
                db.session.add(log)
                update_vm_from_detail(vm, detail)

            vm.last_sync_at = datetime.now()
            vm.last_sync_data = json.dumps(detail, ensure_ascii=False, default=str)
            vm.openstack_project_id = detail.get('project_id')
            db.session.commit()

            flash('同步成功', 'success')
        else:
            vm.openstack_status = 'NOT_FOUND'
            db.session.commit()
            flash('在OpenStack中未找到该云主机', 'warning')

    except Exception as e:
        flash(f'同步失败: {str(e)}', 'error')

    return redirect(url_for('vm_detail', id=id))


@app.route('/vms/<int:id>/extend', methods=['POST'])
def vm_extend(id):
    """延期"""
    vm = VirtualMachine.query.get_or_404(id)

    extend_days = int(request.form.get('extend_days', 90))
    reason = request.form.get('reason', '')

    old_expire = vm.expire_date

    if vm.expire_date:
        vm.expire_date = vm.expire_date + timedelta(days=extend_days)
    else:
        vm.expire_date = datetime.now().date() + timedelta(days=extend_days)

    if vm.status == VirtualMachine.STATUS_VM_PENDING:
        vm.status = VirtualMachine.STATUS_IN_USE

    log = ChangeLog(
        vm_id=vm.id,
        change_type=ChangeLog.TYPE_EXTEND,
        operator=request.form.get('operator', 'admin')
    )
    log.set_change_content({
        'message': f'延期 {extend_days} 天',
        'reason': reason,
        'old_expire_date': old_expire.strftime('%Y-%m-%d') if old_expire else None,
        'new_expire_date': vm.expire_date.strftime('%Y-%m-%d')
    })
    db.session.add(log)
    db.session.commit()

    flash(f'已延期 {extend_days} 天', 'success')
    return redirect(url_for('vm_detail', id=id))


# ============ 回收管理 ============
@app.route('/recycle')
def recycle_list():
    """待回收列表"""
    vm_pending = VirtualMachine.query.filter_by(status=VirtualMachine.STATUS_VM_PENDING).all()
    attached_pending = VirtualMachine.query.filter_by(status=VirtualMachine.STATUS_ATTACHED_PENDING).all()

    return render_template('recycle.html', vm_pending=vm_pending, attached_pending=attached_pending)


@app.route('/vms/<int:id>/recycle', methods=['POST'])
def vm_recycle(id):
    """回收云主机"""
    vm = VirtualMachine.query.get_or_404(id)

    if vm.status not in [VirtualMachine.STATUS_IN_USE, VirtualMachine.STATUS_VM_PENDING]:
        flash('当前状态不允许回收虚机', 'error')
        return redirect(url_for('vm_detail', id=id))

    # 调用OpenStack删除
    success, message = openstack_service.delete_server(vm.uuid)

    old_status = vm.status
    if success:
        vm.status = VirtualMachine.STATUS_VM_PENDING
        vm.vm_deleted_date = datetime.now().date()
        vm.openstack_status = 'DELETED'

        # 更新附属资源状态为待回收
        for resource in vm.attached_resources.filter_by(status=AttachedResource.STATUS_ACTIVE):
            resource.status = AttachedResource.STATUS_PENDING
            resource.pending_at = datetime.now()

    log = ChangeLog(
        vm_id=vm.id,
        change_type=ChangeLog.TYPE_RECYCLE_VM,
        operator=request.form.get('operator', 'admin')
    )
    log.set_change_content({
        'message': '回收云主机',
        'openstack_result': message,
        'success': success,
        'old_status': old_status,
        'new_status': vm.status
    })
    db.session.add(log)
    db.session.commit()

    if success:
        flash(f'云主机回收完成: {message}', 'success')
    else:
        flash(f'云主机回收失败，状态未变更: {message}', 'error')

    return redirect(url_for('vm_detail', id=id))


# ============ 附属资源延期和长期保留 ============
@app.route('/vms/<int:id>/extend-attached', methods=['POST'])
def vm_extend_attached(id):
    """附属资源延期"""
    vm = VirtualMachine.query.get_or_404(id)

    if vm.status not in [VirtualMachine.STATUS_VM_PENDING, VirtualMachine.STATUS_ATTACHED_PENDING]:
        flash('当前状态不支持附属资源延期', 'error')
        return redirect(url_for('vm_detail', id=id))

    extend_days = int(request.form.get('extend_days', 30))
    reason = request.form.get('reason', '')

    old_expire = vm.get_attached_expire_date()

    # 计算新的到期日期
    if vm.attached_expire_date:
        vm.attached_expire_date = vm.attached_expire_date + timedelta(days=extend_days)
    elif vm.vm_deleted_date:
        # 基于原始到期日期延期
        base_date = vm.vm_deleted_date + timedelta(days=Config.ATTACHED_RESOURCE_EXPIRE_DAYS)
        vm.attached_expire_date = base_date + timedelta(days=extend_days)
    else:
        vm.attached_expire_date = datetime.now().date() + timedelta(days=extend_days)

    # 如果当前是附属待回收状态，延期后改回虚机待回收
    if vm.status == VirtualMachine.STATUS_ATTACHED_PENDING:
        vm.status = VirtualMachine.STATUS_VM_PENDING

    log = ChangeLog(
        vm_id=vm.id,
        change_type=ChangeLog.TYPE_EXTEND,
        operator=request.form.get('operator', 'admin')
    )
    log.set_change_content({
        'message': f'附属资源延期 {extend_days} 天',
        'reason': reason,
        'old_expire_date': old_expire.strftime('%Y-%m-%d') if old_expire else None,
        'new_expire_date': vm.attached_expire_date.strftime('%Y-%m-%d')
    })
    db.session.add(log)
    db.session.commit()

    flash(f'附属资源已延期 {extend_days} 天', 'success')
    return redirect(url_for('vm_detail', id=id))


@app.route('/vms/<int:id>/set-attached-long-term', methods=['POST'])
def vm_set_attached_long_term(id):
    """设置附属资源长期保留"""
    vm = VirtualMachine.query.get_or_404(id)

    if vm.status not in [VirtualMachine.STATUS_VM_PENDING, VirtualMachine.STATUS_ATTACHED_PENDING]:
        flash('当前状态不支持设置附属资源长期保留', 'error')
        return redirect(url_for('vm_detail', id=id))

    is_long_term = request.form.get('is_attached_long_term') == 'on'
    reason = request.form.get('attached_long_term_reason', '')

    old_value = vm.is_attached_long_term
    vm.is_attached_long_term = is_long_term
    vm.attached_long_term_reason = reason if is_long_term else ''

    # 如果设置为长期，清除到期日期；如果当前是附属待回收，改回虚机待回收
    if is_long_term:
        vm.attached_expire_date = None
        if vm.status == VirtualMachine.STATUS_ATTACHED_PENDING:
            vm.status = VirtualMachine.STATUS_VM_PENDING

    log = ChangeLog(
        vm_id=vm.id,
        change_type=ChangeLog.TYPE_UPDATE,
        operator=request.form.get('operator', 'admin')
    )
    log.set_change_content({
        'message': '设置附属资源长期保留' if is_long_term else '取消附属资源长期保留',
        'is_attached_long_term': is_long_term,
        'reason': reason
    })
    db.session.add(log)
    db.session.commit()

    if is_long_term:
        flash('附属资源已设置为长期保留', 'success')
    else:
        flash('已取消附属资源长期保留', 'success')

    return redirect(url_for('vm_detail', id=id))

@app.route('/vms/<int:id>/recycle-attached', methods=['POST'])
def vm_recycle_attached(id):
    """回收附属资源"""
    vm = VirtualMachine.query.get_or_404(id)

    # 允许 vm_pending（已删除虚机）或 attached_pending 状态回收附属资源
    if vm.status == VirtualMachine.STATUS_VM_PENDING:
        if not vm.vm_deleted_date:
            flash('请先回收虚机', 'error')
            return redirect(url_for('vm_detail', id=id))
    elif vm.status != VirtualMachine.STATUS_ATTACHED_PENDING:
        flash('当前状态不允许回收附属资源', 'error')
        return redirect(url_for('vm_detail', id=id))

    results = []
    success_count = 0
    fail_count = 0
    successful_resources = set()

    # 从VM记录中获取卷信息
    volumes = vm.get_volumes()
    for vol in volumes:
        vol_id = vol.get('id')
        vol_name = vol.get('name', '')
        if vol_id:
            logger.info(f"Deleting volume: {vol_id} ({vol_name})")
            success, msg = openstack_service.delete_volume(vol_id, force=True)
            results.append({
                'type': 'volume',
                'id': vol_id,
                'name': vol_name,
                'success': success,
                'message': msg
            })
            if success:
                success_count += 1
                successful_resources.add((AttachedResource.TYPE_VOLUME, vol_id))
            else:
                fail_count += 1

    # 从VM记录中获取网络信息（通过MAC删除端口）
    networks = vm.get_networks()
    for net in networks:
        port_id = net.get('port_id') or net.get('id')
        mac = net.get('mac')
        ip = net.get('ip', '')
        network_name = net.get('network_name', '')
        if port_id or mac:
            logger.info(f"Deleting port: {port_id or mac} ({ip})")
            try:
                if port_id:
                    success, msg = openstack_service.delete_port(port_id)
                else:
                    # 兼容早期只记录MAC地址的历史数据
                    port = openstack_service.find_port_by_mac(mac)
                    if port:
                        success, msg = openstack_service.delete_port(port.id)
                    else:
                        success, msg = True, "端口不存在或已删除"
            except Exception as e:
                success, msg = False, f"无法确认端口状态，已取消删除: {str(e)}"
            results.append({
                'type': 'port',
                'id': port_id or mac,
                'name': f"{network_name} - {ip}",
                'success': success,
                'message': msg
            })
            if success:
                success_count += 1
                successful_resources.add((AttachedResource.TYPE_PORT, port_id or mac))
            else:
                fail_count += 1

    # 只更新确认删除成功或确认不存在的附属资源记录。
    for resource in vm.attached_resources.filter(
            AttachedResource.status.in_([AttachedResource.STATUS_ACTIVE, AttachedResource.STATUS_PENDING])
    ):
        if (resource.resource_type, resource.resource_id) in successful_resources:
            resource.status = AttachedResource.STATUS_RECYCLED
            resource.recycled_at = datetime.now()

    remaining_resources = vm.attached_resources.filter(
        AttachedResource.status.in_([AttachedResource.STATUS_ACTIVE, AttachedResource.STATUS_PENDING])
    ).all()
    if fail_count == 0 and remaining_resources:
        for resource in remaining_resources:
            results.append({
                'type': resource.resource_type,
                'id': resource.resource_id,
                'name': resource.resource_name,
                'success': False,
                'message': '资源未在当前同步快照中找到，未执行删除'
            })
        fail_count = len(remaining_resources)

    # 仅当所有删除操作成功时，才更新VM状态为完全回收。
    old_status = vm.status
    if fail_count == 0 and not remaining_resources:
        vm.status = VirtualMachine.STATUS_RECYCLED

    log = ChangeLog(
        vm_id=vm.id,
        change_type=ChangeLog.TYPE_RECYCLE_ATTACHED,
        operator=request.form.get('operator', 'admin')
    )
    log.set_change_content({
        'message': f'回收附属资源（成功: {success_count}, 失败: {fail_count}）',
        'old_status': old_status,
        'new_status': vm.status,
        'results': results
    })
    db.session.add(log)
    db.session.commit()

    if fail_count > 0:
        flash(f'附属资源部分回收失败，状态保持待回收。成功: {success_count}，失败: {fail_count}', 'warning')
    else:
        flash(f'附属资源回收完成，共处理 {success_count} 个资源', 'success')

    return redirect(url_for('vm_detail', id=id))


# ============ API接口 ============
@app.route('/api/openstack/servers')
def api_list_servers():
    """获取OpenStack服务器列表"""
    try:
        servers = openstack_service.list_servers(all_projects=True)
        result = []
        for s in servers:
            result.append({
                'id': s.id,
                'name': s.name,
                'status': s.status,
            })
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/openstack/server/<uuid>')
def api_get_server(uuid):
    """获取服务器详情"""
    try:
        detail = openstack_service.get_server_detail(uuid)
        if detail:
            return jsonify({'success': True, 'data': detail})
        return jsonify({'success': False, 'error': 'Not found'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})



# ============ 变更记录管理 ============
@app.route('/vms/<int:vm_id>/logs/add', methods=['GET', 'POST'])
def log_add(vm_id):
    """添加变更记录"""
    vm = VirtualMachine.query.get_or_404(vm_id)

    if request.method == 'POST':
        change_type = request.form.get('change_type', ChangeLog.TYPE_UPDATE)
        operator = request.form.get('operator', '').strip()
        message = request.form.get('message', '').strip()
        changed_at_str = request.form.get('changed_at', '')

        if not message:
            flash('变更内容不能为空', 'error')
            return redirect(url_for('log_add', vm_id=vm_id))

        log = ChangeLog(
            vm_id=vm_id,
            change_type=change_type,
            operator=operator or 'admin'
        )

        # 解析时间
        if changed_at_str:
            try:
                log.changed_at = datetime.strptime(changed_at_str, '%Y-%m-%dT%H:%M')
            except:
                try:
                    log.changed_at = datetime.strptime(changed_at_str, '%Y-%m-%d')
                except:
                    log.changed_at = datetime.now()

        # 设置变更内容
        detail = request.form.get('detail', '').strip()
        content = {'message': message}
        if detail:
            content['detail'] = detail
        log.set_change_content(content)

        db.session.add(log)
        db.session.commit()

        flash('变更记录添加成功', 'success')
        return redirect(url_for('vm_detail', id=vm_id))

    # 变更类型选项
    change_types = [
        (ChangeLog.TYPE_CREATE, '创建记录'),
        (ChangeLog.TYPE_UPDATE, '信息更新'),
        (ChangeLog.TYPE_SYNC, '同步变更'),
        (ChangeLog.TYPE_STATUS_CHANGE, '状态变更'),
        (ChangeLog.TYPE_RECYCLE_VM, '回收虚机'),
        (ChangeLog.TYPE_RECYCLE_ATTACHED, '回收附属资源'),
        (ChangeLog.TYPE_EXTEND, '延期'),
        ('manual', '手动记录'),
        ('other', '其他'),
    ]

    return render_template('log_add.html', vm=vm, change_types=change_types)


@app.route('/logs/<int:id>/edit', methods=['GET', 'POST'])
def log_edit(id):
    """编辑变更记录"""
    log = ChangeLog.query.get_or_404(id)
    vm = log.virtual_machine

    if request.method == 'POST':
        change_type = request.form.get('change_type', log.change_type)
        operator = request.form.get('operator', '').strip()
        message = request.form.get('message', '').strip()
        changed_at_str = request.form.get('changed_at', '')

        if not message:
            flash('变更内容不能为空', 'error')
            return redirect(url_for('log_edit', id=id))

        log.change_type = change_type
        log.operator = operator or log.operator

        # 解析时间
        if changed_at_str:
            try:
                log.changed_at = datetime.strptime(changed_at_str, '%Y-%m-%dT%H:%M')
            except:
                try:
                    log.changed_at = datetime.strptime(changed_at_str, '%Y-%m-%d')
                except:
                    pass

        # 更新变更内容
        detail = request.form.get('detail', '').strip()
        content = {'message': message}
        if detail:
            content['detail'] = detail

        # 保留原有的changes字段（如果有）
        old_content = log.get_change_content()
        if 'changes' in old_content:
            content['changes'] = old_content['changes']

        log.set_change_content(content)

        db.session.commit()

        flash('变更记录更新成功', 'success')
        return redirect(url_for('vm_detail', id=vm.id))

    # 变更类型选项
    change_types = [
        (ChangeLog.TYPE_CREATE, '创建记录'),
        (ChangeLog.TYPE_UPDATE, '信息更新'),
        (ChangeLog.TYPE_SYNC, '同步变更'),
        (ChangeLog.TYPE_STATUS_CHANGE, '状态变更'),
        (ChangeLog.TYPE_RECYCLE_VM, '回收虚机'),
        (ChangeLog.TYPE_RECYCLE_ATTACHED, '回收附属资源'),
        (ChangeLog.TYPE_EXTEND, '延期'),
        ('manual', '手动记录'),
        ('other', '其他'),
    ]

    return render_template('log_edit.html', log=log, vm=vm, change_types=change_types)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5432, debug=app.config['DEBUG'])
