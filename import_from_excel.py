#!/usr/bin/env python3
"""
Excel数据导入脚本
用于从Excel文件导入云主机历史记录
"""
import os
import sys
from datetime import datetime, timedelta
from openpyxl import load_workbook

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import app
from models import db, Project, VirtualMachine, ChangeLog
from config import Config

# ============ 配置区域（根据你的Excel调整） ============

# Excel文件路径
EXCEL_FILE = 'vm_data.xlsx'

# 工作表名称（None表示使用第一个工作表）
SHEET_NAME = None

# 数据起始行（1表示第一行，通常第1行是表头，数据从第2行开始）
DATA_START_ROW = 2

# 列映射配置（根据你的Excel列调整）
# 格式：'字段名': 列号（A=1, B=2, C=3...）
COLUMN_MAPPING = {
    'uuid': 1,  # A列：UUID（必须）
    'name': 2,  # B列：云主机名称
    'project_name': 3,  # C列：所属项目
    'purpose': 4,  # D列：用途说明
    'applicant': 5,  # E列：申请人
    'ip_addresses': 6,  # F列：IP地址（多个用逗号或换行分隔）
    'flavor': 7,  # G列：规格名称
    'vcpus': 8,  # H列：CPU核数
    'ram': 9,  # I列：内存(MB)
    'disk': 10,  # J列：磁盘(GB)
    'image_name': 11,  # K列：镜像名称
    'host': 12,  # L列：宿主机
    'created_date': 13,  # M列：创建日期
    'expire_date': 14,  # N列：到期日期
    'is_long_term': 15,  # O列：是否长期使用（是/否）
    'long_term_reason': 16,  # P列：长期使用理由
    'status': 17,  # Q列：状态（使用期内/虚机待回收/磁盘网卡待回收/完全回收）
    'volume_info': 18,  # R列：存储卷信息
    'vm_deleted_date': 19,  # S列：虚机删除日期
    'remark': 20,  # T列：备注
}

# 项目列表（如果Excel中的项目不存在，会自动创建）
# 可以预定义项目信息
PREDEFINED_PROJECTS = {
    # '项目名称': {'description': '描述', 'owner': '负责人', 'department': '部门'},
}


# ============ 配置结束 ============


def get_cell_value(row, col_num):
    """获取单元格值"""
    if col_num is None or col_num < 1:
        return None
    try:
        cell = row[col_num - 1]
        if cell.value is None:
            return None
        return str(cell.value).strip()
    except IndexError:
        return None


def parse_date(value):
    """解析日期"""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    try:
        # 尝试多种格式
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日']:
            try:
                return datetime.strptime(str(value), fmt).date()
            except:
                continue
        return None
    except:
        return None


def parse_bool(value):
    """解析布尔值"""
    if not value:
        return False
    return str(value).strip().lower() in ['是', 'yes', 'true', '1', 'y']


def parse_status(value):
    """解析状态"""
    if not value:
        return VirtualMachine.STATUS_IN_USE

    status_map = {
        '使用期内': VirtualMachine.STATUS_IN_USE,
        '虚机待回收': VirtualMachine.STATUS_VM_PENDING,
        '磁盘网卡待回收': VirtualMachine.STATUS_ATTACHED_PENDING,
        '完全回收': VirtualMachine.STATUS_RECYCLED,
        'in_use': VirtualMachine.STATUS_IN_USE,
        'vm_pending': VirtualMachine.STATUS_VM_PENDING,
        'attached_pending': VirtualMachine.STATUS_ATTACHED_PENDING,
        'recycled': VirtualMachine.STATUS_RECYCLED,
    }
    return status_map.get(str(value).strip(), VirtualMachine.STATUS_IN_USE)


def parse_int(value):
    """解析整数"""
    if not value:
        return None
    try:
        return int(float(str(value)))
    except:
        return None


def get_or_create_project(name):
    """获取或创建项目"""
    if not name:
        return None

    project = Project.query.filter_by(name=name).first()
    if not project:
        # 检查预定义项目信息
        project_info = PREDEFINED_PROJECTS.get(name, {})
        project = Project(
            name=name,
            description=project_info.get('description', ''),
            owner=project_info.get('owner', ''),
            department=project_info.get('department', '')
        )
        db.session.add(project)
        db.session.flush()
        print(f"  创建项目: {name}")

    return project


def import_data():
    """导入数据"""
    print(f"开始导入: {EXCEL_FILE}")
    print("=" * 50)

    # 加载Excel
    try:
        wb = load_workbook(EXCEL_FILE)
        if SHEET_NAME:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.active
        print(f"工作表: {ws.title}")
    except Exception as e:
        print(f"加载Excel失败: {e}")
        return

    # 统计
    total = 0
    success = 0
    skip = 0
    fail = 0

    # 遍历数据行
    rows = list(ws.iter_rows(min_row=DATA_START_ROW))
    print(f"共 {len(rows)} 行数据")
    print("-" * 50)

    for row in rows:
        total += 1

        # 获取UUID
        uuid = get_cell_value(row, COLUMN_MAPPING.get('uuid'))
        if not uuid:
            print(f"[{total}] 跳过：UUID为空")
            skip += 1
            continue

        # 检查是否已存在
        existing = VirtualMachine.query.filter_by(uuid=uuid).first()
        if existing:
            print(f"[{total}] 跳过：{uuid[:8]}... 已存在")
            skip += 1
            continue

        try:
            # 创建VM记录
            vm = VirtualMachine(uuid=uuid)

            # 基本信息
            vm.name = get_cell_value(row, COLUMN_MAPPING.get('name'))
            vm.host = get_cell_value(row, COLUMN_MAPPING.get('host'))
            vm.flavor = get_cell_value(row, COLUMN_MAPPING.get('flavor'))
            vm.purpose = get_cell_value(row, COLUMN_MAPPING.get('purpose'))
            vm.applicant = get_cell_value(row, COLUMN_MAPPING.get('applicant'))
            vm.long_term_reason = get_cell_value(row, COLUMN_MAPPING.get('long_term_reason'))

            # 长期使用
            vm.is_long_term = parse_bool(get_cell_value(row, COLUMN_MAPPING.get('is_long_term')))

            # 状态
            vm.status = parse_status(get_cell_value(row, COLUMN_MAPPING.get('status')))
            vm.openstack_status = 'DELETED' if vm.status == VirtualMachine.STATUS_RECYCLED else 'UNKNOWN'

            # 日期
            vm.created_date = parse_date(get_cell_value(row, COLUMN_MAPPING.get('created_date')))
            vm.expire_date = parse_date(get_cell_value(row, COLUMN_MAPPING.get('expire_date')))
            vm.vm_deleted_date = parse_date(get_cell_value(row, COLUMN_MAPPING.get('vm_deleted_date')))

            # 如果没有到期日期，自动计算
            if not vm.expire_date and not vm.is_long_term and vm.created_date:
                vm.expire_date = vm.created_date + timedelta(days=Config.VM_DEFAULT_EXPIRE_DAYS)

            # 规格详情
            vcpus = parse_int(get_cell_value(row, COLUMN_MAPPING.get('vcpus')))
            ram = parse_int(get_cell_value(row, COLUMN_MAPPING.get('ram')))
            disk = parse_int(get_cell_value(row, COLUMN_MAPPING.get('disk')))
            if vcpus or ram or disk:
                import json
                vm.flavor_detail = json.dumps({
                    'vcpus': vcpus,
                    'ram': ram,
                    'disk': disk
                }, ensure_ascii=False)

            # 镜像
            image_name = get_cell_value(row, COLUMN_MAPPING.get('image_name'))
            if image_name:
                import json
                vm.image_info = json.dumps({'name': image_name}, ensure_ascii=False)

            # IP地址
            ip_addresses = get_cell_value(row, COLUMN_MAPPING.get('ip_addresses'))
            if ip_addresses:
                import json
                networks = []
                for ip in ip_addresses.replace(',', '\n').replace('，', '\n').replace(';', '\n').split('\n'):
                    ip = ip.strip()
                    if ip:
                        networks.append({'ip': ip, 'network_name': '', 'mac': ''})
                if networks:
                    vm.networks = json.dumps(networks, ensure_ascii=False)

            # 存储卷
            volume_info = get_cell_value(row, COLUMN_MAPPING.get('volume_info'))
            if volume_info:
                import json
                volumes = []
                for vol in volume_info.replace(',', '\n').replace('，', '\n').split('\n'):
                    vol = vol.strip()
                    if vol:
                        volumes.append({'name': vol, 'id': '', 'size': ''})
                if volumes:
                    vm.volumes = json.dumps(volumes, ensure_ascii=False)

            # 项目
            project_name = get_cell_value(row, COLUMN_MAPPING.get('project_name'))
            if project_name:
                project = get_or_create_project(project_name)
                if project:
                    vm.project_id = project.id

            db.session.add(vm)
            db.session.flush()

            # 创建变更记录
            log = ChangeLog(
                vm_id=vm.id,
                change_type=ChangeLog.TYPE_CREATE,
                operator='excel_import'
            )
            remark = get_cell_value(row, COLUMN_MAPPING.get('remark'))
            log.set_change_content({
                'message': 'Excel导入',
                'remark': remark
            })
            db.session.add(log)

            print(f"[{total}] 成功：{vm.name or uuid[:8]}")
            success += 1

        except Exception as e:
            print(f"[{total}] 失败：{uuid[:8]}... - {e}")
            fail += 1
            db.session.rollback()
            continue

    # 提交
    try:
        db.session.commit()
        print("-" * 50)
        print(f"导入完成！成功: {success}, 跳过: {skip}, 失败: {fail}")
    except Exception as e:
        db.session.rollback()
        print(f"提交失败: {e}")


def main():
    """主函数"""
    # 检查文件
    if not os.path.exists(EXCEL_FILE):
        print(f"文件不存在: {EXCEL_FILE}")
        print("请将Excel文件放在当前目录，或修改 EXCEL_FILE 路径")
        return

    # 在Flask应用上下文中执行
    with app.app_context():
        import_data()


if __name__ == '__main__':
    main()